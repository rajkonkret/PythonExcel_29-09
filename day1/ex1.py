from openpyxl import Workbook, load_workbook

# pip install openpyxl

wb = Workbook()  # tworzyby szablon excel
ws = wb.active  # ustawiamy aktywny arkusz

# komórka w arkuszu A1
ws['A1'] = 42
ws.append([1, 2, 3])  # lista int

# zapianie danych do pliku excel
wb.save('sample.xlsx')
wb.close()

# wczytanie pliku excel
workbook = load_workbook('sample.xlsx')
sheet = workbook.active  # ustawiamy aktywny arkusz
print(sheet)  # <Worksheet "Sheet">, nazwa arkusza w pliku excel

print(sheet['A1'])  # <Cell 'Sheet'.A1>
print(sheet['A1'].value)  # wyświetlenie zawartości dla komórki A1, tu: 42

for row in sheet.iter_rows(min_row=1, max_row=3):  # wiersze
    for cell in row:  # kolumny
        print(cell.value)
# NaN  -> None
