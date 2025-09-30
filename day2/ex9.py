import openpyxl

from openpyxl.styles import Font, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.load_workbook('video2.xlsx')
ws = wb.active

ws = wb['Video Games Sales Data']

# https://htmlcolorcodes.com/
# RGB -> FF0000 -> R=FF, G=00, B=00
ws['A1'].font = Font(color='FF0000', bold=True, size=12)  # red
ws['A2'].font = Font(color='0000FF')  # blue

# ws['A1'].fill = PatternFill('lightVertical', start_color='38e3ff')
ws['A1'].fill = PatternFill('darkTrellis', start_color='38e3ff')
# FILL_NONE = 'none'
# FILL_SOLID = 'solid'
# FILL_PATTERN_DARKDOWN = 'darkDown'
# FILL_PATTERN_DARKGRAY = 'darkGray'
# FILL_PATTERN_DARKGRID = 'darkGrid'
# FILL_PATTERN_DARKHORIZONTAL = 'darkHorizontal'
# FILL_PATTERN_DARKTRELLIS = 'darkTrellis'
# FILL_PATTERN_DARKUP = 'darkUp'
# FILL_PATTERN_DARKVERTICAL = 'darkVertical'
# FILL_PATTERN_GRAY0625 = 'gray0625'
# FILL_PATTERN_GRAY125 = 'gray125'
# FILL_PATTERN_LIGHTDOWN = 'lightDown'
# FILL_PATTERN_LIGHTGRAY = 'lightGray'
# FILL_PATTERN_LIGHTGRID = 'lightGrid'
# FILL_PATTERN_LIGHTHORIZONTAL = 'lightHorizontal'
# FILL_PATTERN_LIGHTTRELLIS = 'lightTrellis'
# FILL_PATTERN_LIGHTUP = 'lightUp'
# FILL_PATTERN_LIGHTVERTICAL = 'lightVertical'
# FILL_PATTERN_MEDIUMGRAY = 'mediumGray'

# ramka
# 'dashDot','dashDotDot', 'dashed','dotted',
#                             'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
#                             'mediumDashed', 'slantDashDot', 'thick', 'thin'
# my_border = Side(border_style='thin', color="000000")  # czarny
my_border = Side(border_style='thick', color="000000")  # czarny
ws['A1'].border = Border(
    top=my_border, left=my_border, right=my_border, bottom=my_border
)

# formatowanie warunkowe - oznaczenie komórek spłęniających określony warunek
fill = PatternFill(
    start_color='90EE90',
    end_color='90EE90',
    fill_type="solid"
)
# {">": "greaterThan", ">=": "greaterThanOrEqual", "<": "lessThan", "<=": "lessThanOrEqual",
#               "=": "equal", "==": "equal", "!=": "notEqual"}
ws.conditional_formatting.add(
    'G2:L16328',
    CellIsRule(operator='lessThan',
               formula=[5],
               fill=fill,
               font=Font(color="FF00FF"))
)

fill_gt = PatternFill(
    start_color='30EE30',
    end_color='30FFFF',
    fill_type="darkUp"
)

ws.conditional_formatting.add(
    'G2:L16328',
    CellIsRule(operator='greaterThan',
               formula=[8],
               fill=fill_gt,
               font=Font(color="FF00FF"))
)

#     "solid",
#     "darkDown",
#     "darkGray",
#     "darkGrid",
#     "darkHorizontal",
#     "darkTrellis",
#     "darkUp",
#     "darkVertical",
#     "gray0625",
#     "gray125",
#     "lightDown",
#     "lightGray",
#     "lightGrid",
#     "lightHorizontal",
#     "lightTrellis",
#     "lightUp",
#     "lightVertical",
#     "mediumGray",
wb.save('video2.xlsx')
wb.close()
