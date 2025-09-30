import openpyxl

wb = openpyxl.load_workbook('video2.xlsx')

ws = wb['vgsales']

ws['P1'] = "Average Sales"
ws['P2'] = '=AVERAGE(L2:L16328)'

wb.save('video2.xlsx')
wb.close()

ws['Q1'] = "Number of populated cells"
ws['Q2'] = "=COUNTA(E2:E16328)"

wb.save('video2.xlsx')
wb.close()

ws['S1'] = "Total Sports Sales"
ws['S2'] = '=COUNTIF(E2:E16328, "sports")'

wb.save('video2.xlsx')
wb.close()

ws['T1'] = "total sum of sports Sales"
ws['T2'] = '=SUMIF(E2:E16328, "Sports", L2:L16328)'

wb.save('video2.xlsx')
wb.close()

ws['U1'] = "Rounded sum of Sports Sales"
ws['U2'] = '=CEILING(T2, 25)'

wb.save('video2.xlsx')
wb.close()

ws['V1'] = "Rounded sum of Sports Sales"
ws['V2'] = '=CEILING(T2, 1)'

wb.save('video2.xlsx')
wb.close()

ws['W1'] = 'Rounded'
ws['W2'] = '=ROUND(T2, 0)'

wb.save('video2.xlsx')
wb.close()

ws['X1'] = "Floor"
ws['X2'] = '=FLOOR(T2, 25)'

wb.save('video2.xlsx')
wb.close()